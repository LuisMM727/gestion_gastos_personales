from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Expense, UserProfile
from .forms import ExpenseForm, UserRegisterForm, UserProfileForm
import datetime
from django.core.paginator import Paginator
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# ==========================================
# BLOQUE 1: GESTIÓN DE USUARIOS Y PERFIL
# ==========================================


# Maneja el registro de nuevos usuarios
def register(request):
    if request.method == "POST":
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, "Cuenta creada exitosamente. Ahora puedes iniciar sesión."
            )
            return redirect("login")
    else:
        form = UserRegisterForm()
    return render(request, "members/register.html", {"form": form})


# Gestiona el perfil del usuario autenticado
@login_required
def profile(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    if request.method == "POST":
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado exitosamente.")
            return redirect("expense_list")
    else:
        form = UserProfileForm(instance=profile)
    return render(request, "members/profile.html", {"form": form})


# Página de inicio: redirige si ya estás logueado o muestra el registro
def home(request):
    if request.user.is_authenticated:
        return redirect("expense_list")
    # Si no está autenticado, mostrar el formulario de registro directamente
    if request.method == "POST":
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, "Cuenta creada exitosamente. Ahora puedes iniciar sesión."
            )
            return redirect("login")
    else:
        form = UserRegisterForm()
    return render(request, "members/register.html", {"form": form})


# ==========================================
# BLOQUE 2: CRUD DE GASTOS (Leer, Crear, Editar, Borrar)
# ==========================================


# LISTAR: Muestra todos los gastos del usuario y hace cálculos financieros
# LISTAR: Muestra los gastos filtrados y recalcula los totales
@login_required
def expense_list(request):
    # 1. Capturar las fechas enviadas desde el formulario de la página (si existen)
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # 2. Empezar con todos los gastos del usuario
    expenses_query = Expense.objects.filter(user=request.user).order_by("-date")

    # 3. APLICAR FILTROS: Si el usuario eligió fechas, filtramos la base de datos
    if start_date:
        expenses_query = expenses_query.filter(date__gte=start_date)
    if end_date:
        expenses_query = expenses_query.filter(date__lte=end_date)

    # 4. Lógica financiera: Se calcula SOLO sobre los registros filtrados
    # Esto asegura que si filtras un mes, los cuadros de arriba sumen solo ese mes
    total_gastado = sum(expense.amount for expense in expenses_query)
    profile = UserProfile.objects.filter(user=request.user).first()
    salario_minimo = profile.salario_minimo if profile else 0
    diferencia = salario_minimo - total_gastado if salario_minimo > 0 else 0

    # 5. Configuración del Paginador
    paginator = Paginator(expenses_query, 7)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "members/expense_list.html",
        {
            "expenses": page_obj,
            "total_gastado": total_gastado,
            "salario_minimo": salario_minimo,
            "diferencia": diferencia,
            # IMPORTANTE: Enviamos las fechas de vuelta al HTML para que no se borren de los cuadritos
            "start_date": start_date or "",
            "end_date": end_date or "",
        },
    )


# CREAR: Procesa el formulario para añadir un nuevo gasto


@login_required
def expense_create(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            messages.success(request, "Gasto creado exitosamente.")
            return redirect("expense_list")
    else:
        form = ExpenseForm(initial={"date": datetime.date.today()})
    return render(request, "members/expense_form.html", {"form": form})


# ACTUALIZAR: Busca un gasto específico por su ID (pk) y lo edita


@login_required
def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk, user=request.user)

    # 1. CAPTURAR EL ESTADO ACTUAL (venga por GET o por POST en la URL)
    page = request.GET.get("page", 1)
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Gasto actualizado exitosamente.")

            # 2. REDIRECCIÓN CONSTRUIDA CON LOS FILTROS CAPTURADOS
            # Esto es lo que evita que el filtro se limpie al guardar
            return redirect(
                f"/expenses/?page={page}&start_date={start_date}&end_date={end_date}"
            )
    else:
        form = ExpenseForm(instance=expense)

    return render(request, "members/expense_form.html", {"form": form})


# ELIMINAR: Borra un registro tras confirmación


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk, user=request.user)
    if request.method == "POST":
        expense.delete()
        messages.success(request, "Gasto eliminado exitosamente.")
        return redirect("expense_list")
    return render(request, "members/expense_confirm_delete.html", {"expense": expense})


# ==========================================
# BLOQUE 3: EXPORTACIÓN DE DATOS
# ==========================================


@login_required
def expense_export_excel(request):
    from openpyxl import Workbook
    from django.http import HttpResponse

    # 1. CAPTURAR LOS FILTROS DE LA URL
    # Obtenemos las fechas si el usuario filtró en la pantalla anterior
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # 2. INICIAR LA CONSULTA BASE
    expenses_query = Expense.objects.filter(user=request.user).order_by("-date")

    # 3. APLICAR FILTROS (Misma lógica que en expense_list)
    if start_date:
        expenses_query = expenses_query.filter(date__gte=start_date)
    if end_date:
        expenses_query = expenses_query.filter(date__lte=end_date)

    # 4. CREACIÓN DEL LIBRO EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos Filtrados"

    # Define los encabezados
    ws["A1"] = "Fecha"
    ws["B1"] = "Descripción"
    ws["C1"] = "Categoría"
    ws["D1"] = "Monto (Gs)"

    # 5. RECORRER LA CONSULTA FILTRADA
    # Ahora 'expenses_query' solo contiene lo que el usuario filtró
    for i, expense in enumerate(expenses_query, start=2):
        ws[f"A{i}"] = expense.date.strftime(
            "%d/%m/%Y"
        )  # Formato más legible (DD/MM/AAAA)
        ws[f"B{i}"] = expense.description
        ws[f"C{i}"] = expense.category
        ws[f"D{i}"] = float(expense.amount)

    # Configuración de la respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Nombre de archivo dinámico si hay fechas
    filename = "gastos.xlsx"
    if start_date or end_date:
        filename = f"gastos_filtrados_{start_date}_a_{end_date}.xlsx"

    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


@login_required
def expense_export_pdf(request):
    from django.http import HttpResponse

    # 1. CAPTURAR FILTROS
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # 2. CONSULTA FILTRADA
    expenses_query = Expense.objects.filter(user=request.user).order_by("-date")
    if start_date:
        expenses_query = expenses_query.filter(date__gte=start_date)
    if end_date:
        expenses_query = expenses_query.filter(date__lte=end_date)

    # 3. CONFIGURAR RESPUESTA HTTP
    response = HttpResponse(content_type="application/pdf")
    filename = "reporte_gastos.pdf"
    if start_date or end_date:
        filename = f"reporte_filtrado_{start_date}_a_{end_date}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # 4. GENERAR EL PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    styles["Title"].alignment = 1
    titulo_texto = "Reporte de Gastos Personales"
    elements.append(Paragraph(titulo_texto, styles["Title"]))

    # 1. DEFINIR ESTILO CENTRADO PARA EL SUBTÍTULO
    estilo_subtitulo = styles["Heading2"]
    estilo_subtitulo.alignment = 1

    # 1. SALTO DE LÍNEA Y FORMATO DE FECHA EN EL SUBTÍTULO
    if start_date and end_date:
        # Convertimos las fechas de AAAA-MM-DD a DD/MM/AAAA para el reporte
        from datetime import datetime

        sd_obj = datetime.strptime(start_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        ed_obj = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d/%m/%Y")

        # Agregamos el periodo como un párrafo nuevo (segunda línea)
        subtitulo = f"Periodo: {sd_obj} al {ed_obj}"
        elements.append(
            Paragraph(subtitulo, styles["Heading2"])
        )  # 'Heading2' lo centra y lo hace un poco más pequeño

    elements.append(Spacer(1, 12))

    # Definir datos para la tabla
    data = [["Fecha", "Descripción", "Categoría", "Monto (Gs)"]]
    for exp in expenses_query:
        data.append(
            [
                exp.date.strftime("%d/%m/%Y"),
                exp.description,
                exp.category,
                f"{exp.amount:,.0f}".replace(",", "."),  # Formato de moneda paraguaya
            ]
        )
    # Estilo de la tabla
    tabla = Table(data, colWidths=[80, 220, 100, 100])
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),  # Descripción a la izquierda
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
            ]
        )
    )

    elements.append(tabla)
    doc.build(elements)
    return response
